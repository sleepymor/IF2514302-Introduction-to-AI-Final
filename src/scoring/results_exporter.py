"""Export benchmark results to Excel format."""

import csv
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from results_analyzer import ResultsAnalyzer
from test_runner import GameResult

try:
    from tqdm import tqdm

    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False


class ResultsExporter:
    """Exports results to Excel and CSV formats."""

    def __init__(self, output_dir: str = None):
        """Initialize exporter.

        Args:
            output_dir: Output directory for Excel files. Defaults to data/results.
        """
        if output_dir is None:
            output_dir = str(Path(__file__).parent.parent.parent / "data" / "results")

        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export_to_excel(self, results: List[GameResult], filename: str = None):
        """Export results to Excel file with multiple sheets.

        Args:
            results: List of GameResult objects.
            filename: Output filename (optional).
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("ERROR: openpyxl not installed. Run: pip install openpyxl")
            print("Falling back to CSV export...")
            self.export_to_csv(results, filename)
            return

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"benchmark_results_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        print("\n[EXPORTING] Creating workbook...")

        # Create workbook with custom sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        analyzer = ResultsAnalyzer(results)

        # Create sheets
        print("[EXPORTING] Creating Summary sheet...")
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, analyzer)

        print("[EXPORTING] Creating Raw Data sheet...")
        ws_raw = wb.create_sheet("Raw Data")
        self._create_raw_data_sheet(ws_raw, analyzer)

        print("[EXPORTING] Creating Per-Seed Stats sheet...")
        ws_per_seed = wb.create_sheet("Per-Seed Stats")
        self._create_per_seed_sheet(ws_per_seed, analyzer)

        # Save workbook
        print("[EXPORTING] Saving Excel file...")
        wb.save(filepath)
        print(f"\n✓ Results exported to: {filepath}")
        return filepath

    def _create_summary_sheet(self, ws, analyzer: ResultsAnalyzer):
        """Create summary statistics sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.title = "Summary"

        # Header styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Get summary stats
        stats = analyzer.get_stats_by_algorithm()

        # Column headers
        headers = [
            "Algorithm",
            "Total Tests",
            "Wins",
            "Win Rate (%)",
            "Deaths by Trap",
            "Death by Trap Rate (%)",
            "Deaths by Enemy",
            "Death by Enemy Rate (%)",
            "Timeouts",
            "Timeout Rate (%)",
            "Errors",
            "Average Moves to Win",
            "Average Turns",
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        # Write data with progress
        data_iterator = (
            tqdm(sorted(stats.items()), desc="  Writing Summary", leave=False)
            if HAS_TQDM
            else sorted(stats.items())
        )

        for row_idx, (algo, stat) in enumerate(data_iterator, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = stat.get(header, "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_raw_data_sheet(self, ws, analyzer: ResultsAnalyzer):
        """Create raw data sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.title = "Raw Data"

        header_fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        raw_data = analyzer.get_raw_data()

        if not raw_data:
            return

        # Write headers
        headers = list(raw_data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data with progress
        data_iterator = (
            tqdm(raw_data, desc="  Writing Raw Data", leave=False)
            if HAS_TQDM
            else raw_data
        )

        for row_idx, row_data in enumerate(data_iterator, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_per_seed_sheet(self, ws, analyzer: ResultsAnalyzer):
        """Create per-seed statistics sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.title = "Per-Seed Stats"

        header_fill = PatternFill(
            start_color="FFC000", end_color="FFC000", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        per_seed_stats = analyzer.get_stats_by_seed()

        if not per_seed_stats:
            return

        # Column headers
        headers = [
            "Seed",
            "Total Tests",
            "Wins",
            "Win Rate (%)",
            "Deaths by Trap",
            "Deaths by Enemy",
            "Timeouts",
            "Avg Moves",
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data with progress
        data_iterator = (
            tqdm(
                sorted(per_seed_stats.items()),
                desc="  Writing Per-Seed Stats",
                leave=False,
            )
            if HAS_TQDM
            else sorted(per_seed_stats.items())
        )

        for row_idx, (seed, stat) in enumerate(data_iterator, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = stat.get(header, "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

    def export_to_csv(self, results: List[GameResult], filename: str = None):
        """Export results to CSV format as fallback.

        Args:
            results: List of GameResult objects.
            filename: Output filename (optional).
        """
        if filename is None:
            filename = "benchmark_results.csv"

        filepath = os.path.join(self.output_dir, filename)

        analyzer = ResultsAnalyzer(results)
        raw_data = analyzer.get_raw_data()

        if not raw_data:
            print("No data to export")
            return

        print("[EXPORTING] Writing CSV files...")

        # Write raw data
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=raw_data[0].keys())
            writer.writeheader()
            writer.writerows(raw_data)

        print(f"\n✓ Results exported to: {filepath}")

        # Write summary stats
        summary_filepath = os.path.join(self.output_dir, "benchmark_summary.csv")
        stats = analyzer.get_stats_by_algorithm()

        with open(summary_filepath, "w", newline="", encoding="utf-8") as f:
            if stats:
                writer = csv.DictWriter(
                    f, fieldnames=stats[list(stats.keys())[0]].keys()
                )
                writer.writeheader()
                for stat in stats.values():
                    writer.writerow(stat)

        print(f"✓ Summary exported to: {summary_filepath}")

        return filepath

    def _create_summary_sheet(self, ws, analyzer: ResultsAnalyzer):
        """Create summary statistics sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.title = "Summary"

        # Header styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Get summary stats
        stats = analyzer.get_stats_by_algorithm()

        # Column headers
        headers = [
            "Algorithm",
            "Total Tests",
            "Wins",
            "Win Rate (%)",
            "Deaths by Trap",
            "Death by Trap Rate (%)",
            "Deaths by Enemy",
            "Death by Enemy Rate (%)",
            "Timeouts",
            "Timeout Rate (%)",
            "Errors",
            "Average Moves to Win",
            "Average Turns",
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        # Write data with progress
        data_iterator = (
            tqdm(sorted(stats.items()), desc="  Writing Summary", leave=False)
            if HAS_TQDM
            else sorted(stats.items())
        )

        for row_idx, (algo, stat) in enumerate(data_iterator, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = stat.get(header, "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_raw_data_sheet(self, ws, analyzer: ResultsAnalyzer):
        """Create raw data sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.title = "Raw Data"

        header_fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        raw_data = analyzer.get_raw_data()

        if not raw_data:
            return

        # Write headers
        headers = list(raw_data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data with progress
        data_iterator = (
            tqdm(raw_data, desc="  Writing Raw Data", leave=False)
            if HAS_TQDM
            else raw_data
        )

        for row_idx, row_data in enumerate(data_iterator, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_per_seed_sheet(self, ws, analyzer: ResultsAnalyzer):
        """Create per-seed statistics sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.title = "Per-Seed Stats"

        header_fill = PatternFill(
            start_color="FFC000", end_color="FFC000", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        per_seed_stats = analyzer.get_stats_by_seed()

        if not per_seed_stats:
            return

        # Column headers
        headers = [
            "Seed",
            "Total Tests",
            "Wins",
            "Win Rate (%)",
            "Deaths by Trap",
            "Deaths by Enemy",
            "Timeouts",
            "Avg Moves",
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data with progress
        data_iterator = (
            tqdm(
                sorted(per_seed_stats.items()),
                desc="  Writing Per-Seed Stats",
                leave=False,
            )
            if HAS_TQDM
            else sorted(per_seed_stats.items())
        )

        for row_idx, (seed, stat) in enumerate(data_iterator, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = stat.get(header, "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

    def export_to_csv(self, results: List[GameResult], filename: str = None):
        """Export results to CSV format as fallback.

        Args:
            results: List of GameResult objects.
            filename: Output filename (optional).
        """
        if filename is None:
            filename = "benchmark_results.csv"

        filepath = os.path.join(self.output_dir, filename)

        analyzer = ResultsAnalyzer(results)
        raw_data = analyzer.get_raw_data()

        if not raw_data:
            print("No data to export")
            return

        print("[EXPORTING] Writing CSV files...")

        # Write raw data
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=raw_data[0].keys())
            writer.writeheader()
            writer.writerows(raw_data)

        print(f"\n✓ Results exported to: {filepath}")

        # Write summary stats
        summary_filepath = os.path.join(self.output_dir, "benchmark_summary.csv")
        stats = analyzer.get_stats_by_algorithm()

        with open(summary_filepath, "w", newline="", encoding="utf-8") as f:
            if stats:
                writer = csv.DictWriter(
                    f, fieldnames=stats[list(stats.keys())[0]].keys()
                )
                writer.writeheader()
                for stat in stats.values():
                    writer.writerow(stat)

        print(f"✓ Summary exported to: {summary_filepath}")

        return filepath
