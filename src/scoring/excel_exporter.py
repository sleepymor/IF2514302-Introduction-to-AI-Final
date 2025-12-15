"""Excel export functionality for benchmark results.

Exports comprehensive results to Excel with multiple sheets:
- Summary: Overall results by algorithm
- Per Seed: Results per algorithm-seed combination
- Episode Details: Individual episode data
"""

from datetime import datetime
from pathlib import Path
from typing import Dict

from benchmark_data import BenchmarkConfig, BenchmarkResults


class ExcelExporter:
    """Exports benchmark results to Excel with multiple sheets."""

    OUTPUT_DIR = None  # Lazily initialized

    @classmethod
    def _ensure_output_dir(cls) -> Path:
        """Ensure output directory exists.

        Returns:
            Path to output directory
        """
        if cls.OUTPUT_DIR is None:
            cls.OUTPUT_DIR = Path(__file__).parent.parent.parent / "data" / "results"
            cls.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        return cls.OUTPUT_DIR

    @classmethod
    def export(
        cls, config: BenchmarkConfig, results: Dict[str, BenchmarkResults]
    ) -> None:
        """Export comprehensive results to Excel with multiple sheets.

        Args:
            config: BenchmarkConfig used for evaluation
            results: Algorithm name -> BenchmarkResults mapping
        """
        try:
            import openpyxl
        except ImportError:
            print(
                "[WARNING] openpyxl not installed. Skipping Excel export. "
                "Install with: pip install openpyxl"
            )
            return

        output_dir = cls._ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = output_dir / f"benchmark_results_{timestamp}.xlsx"

        print("[EXPORTING] Writing comprehensive results to Excel...")

        total_episodes = len(config.environment_seeds) * config.episodes_per_seed
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Create sheets
        cls._create_summary_sheet(wb, config, results, total_episodes)
        cls._create_per_seed_sheet(wb, config, results)
        cls._create_episode_details_sheet(wb, config, results)

        wb.save(filepath)
        print(f"✓ Excel workbook saved to: {filepath}\n")
        print("  Sheets included:")
        print("    1. Summary - Overall results by algorithm")
        print("    2. Per Seed - Results for each algorithm-seed combination")
        print("    3. Episode Details - Seed-by-seed, episode-by-episode results\n")

    @classmethod
    def _create_summary_sheet(
        cls,
        wb,
        config: BenchmarkConfig,
        results: Dict[str, BenchmarkResults],
        total_episodes: int,
    ) -> None:
        """Create Summary sheet in workbook.

        Args:
            wb: openpyxl Workbook
            config: BenchmarkConfig used for evaluation
            results: Algorithm name -> BenchmarkResults mapping
            total_episodes: Total episodes per algorithm
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Summary", 0)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        headers = [
            "Algorithm",
            "Total Episodes",
            "Seeds Tested",
            "Episodes per Seed",
            "Max Moves",
            "Goal Reached",
            "Goal Rate (%)",
            "Hit Trap",
            "Trap Rate (%)",
            "Caught by Enemy",
            "Caught Rate (%)",
            "Timeout",
            "Timeout Rate (%)",
            "Win Rate (%)",
            "Avg Moves",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row, algo in enumerate(config.algorithms, 2):
            algo_results = results[algo]
            win_rate = algo_results.get_win_rate(total_episodes)
            avg_moves = algo_results.get_average_moves()

            ws.cell(row=row, column=1).value = algo
            ws.cell(row=row, column=2).value = total_episodes
            ws.cell(row=row, column=3).value = len(config.environment_seeds)
            ws.cell(row=row, column=4).value = config.episodes_per_seed
            ws.cell(row=row, column=5).value = config.max_moves_per_episode
            ws.cell(row=row, column=6).value = algo_results.outcomes.get("goal", 0)
            ws.cell(row=row, column=7).value = (
                algo_results.outcomes.get("goal", 0) / total_episodes * 100
            )
            ws.cell(row=row, column=8).value = algo_results.outcomes.get("trap", 0)
            ws.cell(row=row, column=9).value = (
                algo_results.outcomes.get("trap", 0) / total_episodes * 100
            )
            ws.cell(row=row, column=10).value = algo_results.outcomes.get("caught", 0)
            ws.cell(row=row, column=11).value = (
                algo_results.outcomes.get("caught", 0) / total_episodes * 100
            )
            ws.cell(row=row, column=12).value = algo_results.outcomes.get("timeout", 0)
            ws.cell(row=row, column=13).value = (
                algo_results.outcomes.get("timeout", 0) / total_episodes * 100
            )
            ws.cell(row=row, column=14).value = win_rate * 100
            ws.cell(row=row, column=15).value = avg_moves

            # Format
            for col in [7, 9, 11, 13, 14, 15]:
                ws.cell(row=row, column=col).number_format = "0.0"

        # Column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    @classmethod
    def _create_per_seed_sheet(
        cls, wb, config: BenchmarkConfig, results: Dict[str, BenchmarkResults]
    ) -> None:
        """Create Per Seed sheet in workbook.

        Args:
            wb: openpyxl Workbook
            config: BenchmarkConfig used for evaluation
            results: Algorithm name -> BenchmarkResults mapping
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Per Seed", 1)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        headers = [
            "Algorithm",
            "Seed",
            "Episodes",
            "Goal Reached",
            "Goal Rate (%)",
            "Hit Trap",
            "Trap Rate (%)",
            "Caught",
            "Caught Rate (%)",
            "Timeout",
            "Timeout Rate (%)",
            "Win Rate (%)",
            "Avg Moves",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        row = 2
        for algo in config.algorithms:
            algo_results = results[algo]
            for seed in config.environment_seeds:
                if seed in algo_results.outcomes_by_seed:
                    seed_outcomes = algo_results.outcomes_by_seed[seed]
                    total = config.episodes_per_seed
                    win_rate = (
                        seed_outcomes.get("goal", 0) / total * 100 if total > 0 else 0
                    )
                    avg_moves = algo_results.get_seed_average_moves(seed)

                    ws.cell(row=row, column=1).value = algo
                    ws.cell(row=row, column=2).value = seed
                    ws.cell(row=row, column=3).value = total
                    ws.cell(row=row, column=4).value = seed_outcomes.get("goal", 0)
                    ws.cell(row=row, column=5).value = (
                        seed_outcomes.get("goal", 0) / total * 100 if total > 0 else 0
                    )
                    ws.cell(row=row, column=6).value = seed_outcomes.get("trap", 0)
                    ws.cell(row=row, column=7).value = (
                        seed_outcomes.get("trap", 0) / total * 100 if total > 0 else 0
                    )
                    ws.cell(row=row, column=8).value = seed_outcomes.get("caught", 0)
                    ws.cell(row=row, column=9).value = (
                        seed_outcomes.get("caught", 0) / total * 100 if total > 0 else 0
                    )
                    ws.cell(row=row, column=10).value = seed_outcomes.get("timeout", 0)
                    ws.cell(row=row, column=11).value = (
                        seed_outcomes.get("timeout", 0) / total * 100
                        if total > 0
                        else 0
                    )
                    ws.cell(row=row, column=12).value = win_rate
                    ws.cell(row=row, column=13).value = avg_moves

                    # Format
                    for col in [5, 7, 9, 11, 12, 13]:
                        ws.cell(row=row, column=col).number_format = "0.0"

                    row += 1

        # Column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    @classmethod
    def _create_episode_details_sheet(
        cls, wb, config: BenchmarkConfig, results: Dict[str, BenchmarkResults]
    ) -> None:
        """Create Episode Details sheet in workbook.

        Args:
            wb: openpyxl Workbook
            config: BenchmarkConfig used for evaluation
            results: Algorithm name -> BenchmarkResults mapping
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Episode Details", 2)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Algorithm", "Seed", "Episode", "Outcome", "Moves"]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        row = 2
        for algo in config.algorithms:
            algo_results = results[algo]
            sorted_episodes = sorted(
                algo_results.episode_details, key=lambda x: (x["seed"], x["episode"])
            )
            for ep_detail in sorted_episodes:
                ws.cell(row=row, column=1).value = algo
                ws.cell(row=row, column=2).value = ep_detail["seed"]
                ws.cell(row=row, column=3).value = ep_detail["episode"]
                ws.cell(row=row, column=4).value = ep_detail["outcome"]
                ws.cell(row=row, column=5).value = ep_detail["moves"]
                row += 1

        # Column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
